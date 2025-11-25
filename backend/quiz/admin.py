# quiz/admin.py
from django.contrib import admin
from django.contrib.auth.models import User
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.hashers import make_password
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.conf import settings
from django.core.files import File
from import_export import resources, fields
from import_export.admin import ImportExportModelAdmin
from import_export.widgets import ForeignKeyWidget
import os
import tempfile

from .models import Subtest, Soal, HasilTryout


@admin.register(Subtest)
class SubtestAdmin(admin.ModelAdmin):
    list_display = ('code', 'nama_subtest', 'jumlah_soal', 'durasi_menit')
    list_editable = ('jumlah_soal', 'durasi_menit')
    search_fields = ('code', 'nama_subtest')


class SoalResource(resources.ModelResource):
    """
    Resource untuk impor/ekspor bank soal.

    Disesuaikan dengan format Excel yang kamu tunjukkan:

    Kode Subtes | SOAL | A | B | C | D | E | KUNCI
    """

    # mapping kolom Excel → field model
    subtest_code = fields.Field(
        column_name="Kode Subtes",
        attribute="subtest",
        widget=ForeignKeyWidget(Subtest, "code"),
    )
    soal_text = fields.Field(column_name="SOAL", attribute="soal_text")
    option_a = fields.Field(column_name="A", attribute="option_a")
    option_b = fields.Field(column_name="B", attribute="option_b")
    option_c = fields.Field(column_name="C", attribute="option_c")
    option_d = fields.Field(column_name="D", attribute="option_d")
    option_e = fields.Field(column_name="E", attribute="option_e")
    correct_answer = fields.Field(column_name="KUNCI", attribute="correct_answer")

    class Meta:
        model = Soal
        # gunakan nama field resource (bukan nama field model) untuk urutan ekspor
        fields = (
            "subtest_code",
            "soal_text",
            "option_a",
            "option_b",
            "option_c",
            "option_d",
            "option_e",
            "correct_answer",
        )
        export_order = fields
        import_id_fields = ()  # tidak ada unique field untuk skip duplicate
        skip_unchanged = False

    def before_import_row(self, row, **kwargs):
        """Validasi & normalisasi data sebelum import."""
        # pastikan KUNCI adalah A/B/C/D/E
        kunci = str(row.get("KUNCI", "")).strip().upper()
        if kunci not in ["A", "B", "C", "D", "E"]:
            raise ValueError(f"KUNCI harus A/B/C/D/E, ditemukan: {kunci}")

    def import_row(self, row, instance_loader, using_transactions=True, dry_run=False, **kwargs):
        """Override untuk handle error dengan lebih baik."""
        try:
            return super().import_row(row, instance_loader, using_transactions, dry_run, **kwargs)
        except Exception as e:
            # log error yang lebih jelas
            row_num = kwargs.get("row_number", "?")
            raise ValueError(f"Error di baris {row_num}: {str(e)}")


@admin.register(Soal)
class SoalAdmin(ImportExportModelAdmin):
    resource_class = SoalResource
    list_display = ('subtest', 'soal_text_preview', 'has_image', 'correct_answer', 'created_at')
    list_filter = ('subtest__code',)
    search_fields = ('soal_text',)
    fieldsets = (
        ('Informasi Subtest', {
            'fields': ('subtest',)
        }),
        ('Soal', {
            'fields': ('soal_text', 'soal_image'),
            'description': 'Isi teks soal atau upload gambar soal. Jika menggunakan gambar, teks soal bisa dikosongkan.'
        }),
        ('Opsi Jawaban', {
            'fields': ('option_a', 'option_b', 'option_c', 'option_d', 'option_e')
        }),
        ('Jawaban Benar', {
            'fields': ('correct_answer',)
        }),
        ('Metadata', {
            'fields': ('created_at',),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ('created_at',)
    
    def soal_text_preview(self, obj):
        """Preview teks soal (maksimal 100 karakter)."""
        if obj.soal_text:
            return obj.soal_text[:100] + '...' if len(obj.soal_text) > 100 else obj.soal_text
        return '(Gambar saja)'
    soal_text_preview.short_description = 'Teks Soal'
    
    def has_image(self, obj):
        """Indikator apakah soal memiliki gambar."""
        return bool(obj.soal_image)
    has_image.short_description = 'Gambar'
    has_image.boolean = True
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel_view), name='quiz_soal_upload_excel'),
        ]
        return custom_urls + urls
    
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['upload_excel_url'] = 'admin:quiz_soal_upload_excel'
        return super().changelist_view(request, extra_context=extra_context)
    
    def upload_excel_view(self, request):
        """Halaman upload Excel untuk import bank soal."""
        if request.method == 'POST':
            if 'file' not in request.FILES:
                messages.error(request, 'File Excel tidak ditemukan.')
                return redirect('admin:quiz_soal_upload_excel')
            
            excel_file = request.FILES['file']
            
            # validasi ekstensi
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'File harus berformat .xlsx atau .xls')
                return redirect('admin:quiz_soal_upload_excel')
            
            # proses file Excel langsung
            try:
                # simpan file sementara
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    for chunk in excel_file.chunks():
                        tmp_file.write(chunk)
                    tmp_path = tmp_file.name
                
                # baca file dengan openpyxl dan proses langsung
                try:
                    from openpyxl import load_workbook
                    
                    wb = load_workbook(tmp_path, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # baca header
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    expected_headers = ["Kode Subtes", "SOAL", "A", "B", "C", "D", "E", "KUNCI"]
                    has_gambar_column = len(header_row) > 8 and str(header_row[8]).strip().lower() in ['gambar', 'gambar_soal', 'image']
                    
                    # validasi header (8 kolom wajib)
                    header_lower = [str(h).strip().lower() if h else "" for h in header_row[:8]]
                    expected_lower = [h.lower() for h in expected_headers]
                    
                    if header_lower[:8] != expected_lower:
                        messages.error(
                            request,
                            f'Header tidak sesuai. Diharapkan: {", ".join(expected_headers)}. '
                            f'Ditemukan: {", ".join([str(h) for h in header_row[:8]])}'
                        )
                        wb.close()
                        os.unlink(tmp_path)
                        return redirect('admin:quiz_soal_upload_excel')
                    
                    created = 0
                    errors = []
                    
                    # proses setiap baris data
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        # skip baris kosong
                        if not any(row[:8]):
                            continue
                        
                        try:
                            kode_subtes = str(row[0]).strip() if row[0] else ""
                            soal_text = str(row[1]).strip() if row[1] else ""
                            option_a = str(row[2]).strip() if row[2] else ""
                            option_b = str(row[3]).strip() if row[3] else ""
                            option_c = str(row[4]).strip() if row[4] else ""
                            option_d = str(row[5]).strip() if row[5] else ""
                            option_e = str(row[6]).strip() if row[6] else ""
                            kunci = str(row[7]).strip().upper() if row[7] else ""
                            gambar_path = str(row[8]).strip() if has_gambar_column and len(row) > 8 and row[8] else ""
                            
                            # validasi
                            if not kode_subtes:
                                errors.append(f"Baris {row_num}: Kode Subtes kosong")
                                continue
                            
                            # SOAL atau GAMBAR harus ada salah satu
                            if not soal_text and not gambar_path:
                                errors.append(f"Baris {row_num}: SOAL atau GAMBAR harus diisi (minimal salah satu)")
                                continue
                            
                            if kunci not in ["A", "B", "C", "D", "E"]:
                                errors.append(f"Baris {row_num}: KUNCI harus A/B/C/D/E, ditemukan: {kunci}")
                                continue
                            
                            # cari subtest
                            try:
                                subtest = Subtest.objects.get(code=kode_subtes)
                            except Subtest.DoesNotExist:
                                errors.append(f"Baris {row_num}: Kode Subtes '{kode_subtes}' tidak ditemukan")
                                continue
                            
                            # Handle gambar
                            soal_image = None
                            if gambar_path:
                                from django.core.files import File
                                from django.core.files.images import ImageFile
                                from urllib.parse import urlparse
                                import requests
                                
                                # Cek apakah path adalah URL
                                parsed = urlparse(gambar_path)
                                if parsed.scheme in ['http', 'https']:
                                    # Download dari URL
                                    try:
                                        response = requests.get(gambar_path, timeout=10)
                                        response.raise_for_status()
                                        # Simpan ke temporary file
                                        img_ext = os.path.splitext(parsed.path)[1] or '.jpg'
                                        with tempfile.NamedTemporaryFile(delete=False, suffix=img_ext) as img_tmp:
                                            img_tmp.write(response.content)
                                            img_tmp_path = img_tmp.name
                                        
                                        # Buka file dan simpan ke model
                                        with open(img_tmp_path, 'rb') as img_file:
                                            soal_image = File(img_file, name=os.path.basename(parsed.path) or f'image_{row_num}{img_ext}')
                                        os.unlink(img_tmp_path)
                                    except Exception as e:
                                        errors.append(f"Baris {row_num}: Gagal download gambar dari URL: {str(e)}")
                                else:
                                    # Path file lokal (relatif ke media/soal_images/)
                                    # Admin harus upload file ke folder media/soal_images/ terlebih dahulu
                                    full_path = os.path.join(settings.MEDIA_ROOT, 'soal_images', gambar_path)
                                    if os.path.exists(full_path):
                                        with open(full_path, 'rb') as img_file:
                                            soal_image = File(img_file, name=os.path.basename(gambar_path))
                                    else:
                                        errors.append(f"Baris {row_num}: File gambar tidak ditemukan: {gambar_path}")
                            
                            # buat soal baru
                            soal = Soal.objects.create(
                                subtest=subtest,
                                soal_text=soal_text or "",  # Bisa kosong jika hanya gambar
                                option_a=option_a,
                                option_b=option_b,
                                option_c=option_c,
                                option_d=option_d,
                                option_e=option_e,
                                correct_answer=kunci,
                            )
                            
                            # Set gambar jika ada
                            if soal_image:
                                soal.soal_image = soal_image
                                soal.save()
                            
                            created += 1
                            
                        except Exception as e:
                            errors.append(f"Baris {row_num}: {str(e)}")
                    
                    wb.close()
                    os.unlink(tmp_path)
                    
                    # tampilkan hasil
                    if created > 0:
                        messages.success(request, f'Berhasil mengimport {created} soal.')
                    if errors:
                        error_msg = f'Ada {len(errors)} error. ' + '; '.join(errors[:5])
                        if len(errors) > 5:
                            error_msg += f' ... dan {len(errors) - 5} error lainnya.'
                        messages.warning(request, error_msg)
                    
                    return redirect('admin:quiz_soal_changelist')
                    
                except ImportError:
                    messages.error(request, 'openpyxl belum terinstall. Install dengan: pip install openpyxl')
                    os.unlink(tmp_path)
                    return redirect('admin:quiz_soal_upload_excel')
                except Exception as e:
                    os.unlink(tmp_path)
                    messages.error(request, f'Gagal membaca file Excel: {str(e)}')
                    return redirect('admin:quiz_soal_upload_excel')
                    
            except Exception as e:
                messages.error(request, f'Error: {str(e)}')
                return redirect('admin:quiz_soal_upload_excel')
        
        # GET request - tampilkan form
        context = {
            **self.admin_site.each_context(request),
            'title': 'Import Bank Soal dari Excel',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/quiz/soal/upload_excel.html', context)


# === Import user dari Excel/CSV ===

class UserResource(resources.ModelResource):
    """
    Resource untuk impor User dari file (Excel/CSV).

    Contoh header kolom:
    username,password,first_name,last_name,email,is_staff

    Password diisi plain text di file, nanti otomatis di-hash sebelum disimpan.
    """

    class Meta:
        model = User
        # kita hanya butuh kolom username & password dari Excel
        # kolom lain akan diisi nilai default
        fields = (
            'username',
            'password',
        )
        import_id_fields = ('username',)

    def before_import_row(self, row, **kwargs):
        # Expect header minimal: username,password
        # password masih plain text → kita hash dulu
        raw_password = row.get('password')
        if raw_password and not str(raw_password).startswith('pbkdf2_'):
            row['password'] = make_password(str(raw_password))



# override admin User bawaan supaya ada tombol Import/Export
admin.site.unregister(User)

@admin.register(User)
class UserAdmin(ImportExportModelAdmin, BaseUserAdmin):
    resource_class = UserResource


@admin.register(HasilTryout)
class HasilTryoutAdmin(admin.ModelAdmin):
    """
    Admin untuk melihat hasil tryout siswa.
    Admin bisa melihat jawaban yang dipilih siswa untuk setiap soal.
    """
    list_display = (
        'user', 
        'subtest', 
        'batch_id', 
        'skor_display', 
        'jumlah_benar', 
        'jumlah_salah', 
        'jumlah_kosong',
        'waktu_selesai',
        'durasi_display'
    )
    list_filter = ('subtest__code', 'batch_id', 'waktu_selesai')
    search_fields = ('user__username', 'user__first_name', 'user__last_name', 'subtest__code')
    readonly_fields = (
        'user', 
        'subtest', 
        'batch_id', 
        'jawaban_display', 
        'jumlah_benar', 
        'jumlah_salah', 
        'jumlah_kosong', 
        'skor',
        'waktu_mulai', 
        'waktu_selesai', 
        'durasi_detik',
        'created_at',
        'updated_at'
    )
    fieldsets = (
        ('Informasi User & Subtest', {
            'fields': ('user', 'subtest', 'batch_id')
        }),
        ('Hasil Pengerjaan', {
            'fields': ('skor', 'jumlah_benar', 'jumlah_salah', 'jumlah_kosong')
        }),
        ('Jawaban Siswa', {
            'fields': ('jawaban_display',),
            'description': 'Jawaban yang dipilih siswa untuk setiap soal. Format: Soal ID → Jawaban Siswa (Benar/Salah)'
        }),
        ('Waktu & Durasi', {
            'fields': ('waktu_mulai', 'waktu_selesai', 'durasi_detik')
        }),
        ('Metadata', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def skor_display(self, obj):
        """Tampilkan skor dengan format yang lebih jelas."""
        if obj.skor is None:
            return "0.00%"
        return f"{float(obj.skor):.2f}%"
    skor_display.short_description = 'Skor'
    skor_display.admin_order_field = 'skor'
    
    def durasi_display(self, obj):
        """Tampilkan durasi dalam format menit:detik."""
        if not obj.durasi_detik:
            return "-"
        menit = obj.durasi_detik // 60
        detik = obj.durasi_detik % 60
        return f"{menit}:{detik:02d}"
    durasi_display.short_description = 'Durasi'
    
    def jawaban_display(self, obj):
        """Tampilkan jawaban siswa dengan format yang mudah dibaca."""
        if not obj.jawaban or len(obj.jawaban) == 0:
            return "Tidak ada jawaban yang tersimpan."
        
        # Ambil semua soal untuk subtest ini
        from .models import Soal
        soal_list = Soal.objects.filter(subtest=obj.subtest).order_by('id')
        
        # Buat mapping soal_id -> soal object
        soal_dict = {str(soal.id): soal for soal in soal_list}
        
        html_parts = ['<div style="max-height: 600px; overflow-y: auto;">']
        html_parts.append('<table style="width: 100%; border-collapse: collapse; font-size: 12px;">')
        html_parts.append('<thead><tr style="background-color: #f0f0f0;">')
        html_parts.append('<th style="padding: 8px; border: 1px solid #ddd; text-align: center;">No</th>')
        html_parts.append('<th style="padding: 8px; border: 1px solid #ddd; text-align: center;">Soal ID</th>')
        html_parts.append('<th style="padding: 8px; border: 1px solid #ddd;">Jawaban Siswa</th>')
        html_parts.append('<th style="padding: 8px; border: 1px solid #ddd;">Jawaban Benar</th>')
        html_parts.append('<th style="padding: 8px; border: 1px solid #ddd; text-align: center;">Status</th>')
        html_parts.append('</tr></thead><tbody>')
        
        # Tampilkan semua soal (termasuk yang tidak dijawab)
        for idx, soal in enumerate(soal_list, start=1):
            soal_id_str = str(soal.id)
            jawaban_user = obj.jawaban.get(soal_id_str, "").strip().upper()
            jawaban_benar = soal.correct_answer.strip().upper()
            
            # Tentukan status
            if not jawaban_user:
                status = '<span style="color: #999;">Kosong</span>'
                bg_color = '#f9f9f9'
            elif jawaban_user == jawaban_benar:
                status = '<span style="color: #28a745; font-weight: bold;">✓ Benar</span>'
                bg_color = '#d4edda'
            else:
                status = '<span style="color: #dc3545; font-weight: bold;">✗ Salah</span>'
                bg_color = '#f8d7da'
            
            html_parts.append(f'<tr style="background-color: {bg_color};">')
            html_parts.append(f'<td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{idx}</td>')
            html_parts.append(f'<td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{soal_id_str}</td>')
            html_parts.append(f'<td style="padding: 8px; border: 1px solid #ddd; font-weight: bold;">{jawaban_user if jawaban_user else "-"}</td>')
            html_parts.append(f'<td style="padding: 8px; border: 1px solid #ddd;">{jawaban_benar}</td>')
            html_parts.append(f'<td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{status}</td>')
            html_parts.append('</tr>')
        
        html_parts.append('</tbody></table>')
        html_parts.append('</div>')
        
        return mark_safe(''.join(html_parts))
    jawaban_display.short_description = 'Detail Jawaban'
    
    def has_add_permission(self, request):
        """Tidak bisa menambah HasilTryout manual dari admin."""
        return False
    
    def has_change_permission(self, request, obj=None):
        """Tidak bisa mengubah HasilTryout dari admin (read-only)."""
        return False
    
    def has_delete_permission(self, request, obj=None):
        """Bisa menghapus jika diperlukan."""
        return request.user.is_superuser
