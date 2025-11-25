from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAdminUser
from django.contrib.auth import authenticate
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Count, Avg, Q
from datetime import datetime
from .models import Subtest, Soal, HasilTryout

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


@api_view(['GET'])
def subtests_list(request):
    """
    GET /api/subtests/
    Returns list of all subtests with their details.
    """
    subtests = Subtest.objects.all().order_by('code')
    data = []
    for subtest in subtests:
        data.append({
            'id': subtest.code.lower(),
            'code': subtest.code,
            'title': subtest.nama_subtest,
            'description': f"Subtest {subtest.nama_subtest} untuk UTBK 2025",
            'duration': float(subtest.durasi_menit),
            'questionCount': subtest.jumlah_soal,
        })
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
def subtest_questions(request, code):
    """
    GET /api/subtests/{code}/questions/
    Returns all questions for a specific subtest.
    """
    try:
        subtest = Subtest.objects.get(code=code.upper())
    except Subtest.DoesNotExist:
        return Response(
            {'error': f'Subtest with code {code} not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    questions = Soal.objects.filter(subtest=subtest).order_by('id')
    data = []
    for idx, soal in enumerate(questions, start=1):
        # URL gambar jika ada
        image_url = None
        if soal.soal_image:
            request_scheme = request.scheme if hasattr(request, 'scheme') else 'http'
            request_host = request.get_host() if hasattr(request, 'get_host') else 'localhost:8000'
            image_url = f"{request_scheme}://{request_host}{soal.soal_image.url}"
        
        data.append({
            'id': f"{subtest.code.lower()}-{idx}",
            'soal_id': soal.id,  # ID sebenarnya dari database
            'subtestId': subtest.code.lower(),
            'question': soal.soal_text,
            'question_image': image_url,  # URL gambar soal (jika ada)
            'options': [
                {'key': 'A', 'text': soal.option_a},
                {'key': 'B', 'text': soal.option_b},
                {'key': 'C', 'text': soal.option_c},
                {'key': 'D', 'text': soal.option_d},
                {'key': 'E', 'text': soal.option_e},
            ],
            'answer': soal.correct_answer,
        })
    return Response(data, status=status.HTTP_200_OK)


@api_view(['POST'])
def login_view(request):
    """
    POST /api/auth/login/

    Body: { "username": "...", "password": "..." }
    Validasi user menggunakan database Django default.
    """
    username = request.data.get('username')
    password = request.data.get('password')

    if not username or not password:
        return Response(
            {'error': 'Username & password wajib'},
            status=status.HTTP_400_BAD_REQUEST
        )

    user = authenticate(request, username=username, password=password)
    if user is None:
        return Response(
            {'error': 'Username atau password salah'},
            status=status.HTTP_401_UNAUTHORIZED
        )

    # Bentuk payload minimal yang akan dipakai frontend
    name = user.get_full_name() or user.username
    role = 'admin' if user.is_staff else 'student'

    return Response(
        {
            'ok': True,
            'user': {
                'username': user.username,
                'name': name,
                'role': role,
            },
        },
        status=status.HTTP_200_OK,
    )


@api_view(['POST'])
@csrf_exempt
def import_soal_excel(request):
    """
    POST /api/import-soal-excel/
    
    Import bank soal dari file Excel.
    Format: Kode Subtes | SOAL | A | B | C | D | E | KUNCI
    
    Body: multipart/form-data dengan field 'file' (file Excel .xlsx)
    """
    if load_workbook is None:
        return Response(
            {'error': 'openpyxl belum terinstall. Install dengan: pip install openpyxl'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    
    if 'file' not in request.FILES:
        return Response(
            {'error': 'File Excel tidak ditemukan. Kirim dengan field "file".'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    excel_file = request.FILES['file']
    
    # validasi ekstensi
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return Response(
            {'error': 'File harus berformat .xlsx atau .xls'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        # baca file Excel dengan openpyxl
        wb = load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.active
        
        # baca header
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        expected_headers = ["Kode Subtes", "SOAL", "A", "B", "C", "D", "E", "KUNCI"]
        
        # Cek apakah ada kolom gambar (opsional, kolom ke-9)
        has_gambar_column = len(header_row) > 8 and str(header_row[8]).strip().lower() in ['gambar', 'gambar_soal', 'image', 'gambar soal']
        
        # validasi header (case-insensitive)
        header_lower = [str(h).strip().lower() if h else "" for h in header_row[:8]]
        expected_lower = [h.lower() for h in expected_headers]
        
        if header_lower[:8] != expected_lower:
            return Response(
                {
                    'error': f'Header tidak sesuai. Diharapkan: {", ".join(expected_headers)}',
                    'found': header_lower[:8]
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created = 0
        errors = []
        
        # proses setiap baris data
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # skip baris kosong
            if not any(row[:8]):
                continue
            
            try:
                kode_subtes = str(row[0]).strip() if row[0] else ""
                soal_text = str(row[1]).strip() if row[1] else ""
                option_a = str(row[2]).strip() if row[2] else ""
                option_b = str(row[3]).strip() if row[3] else ""
                option_c = str(row[4]).strip() if row[4] else ""
                option_d = str(row[5]).strip() if row[5] else ""
                option_e = str(row[6]).strip() if row[6] else ""
                kunci = str(row[7]).strip().upper() if row[7] else ""
                gambar_path = str(row[8]).strip() if has_gambar_column and len(row) > 8 and row[8] else ""
                
                # validasi
                if not kode_subtes:
                    errors.append(f"Baris {row_num}: Kode Subtes kosong")
                    continue
                
                # SOAL atau GAMBAR harus ada salah satu
                if not soal_text and not gambar_path:
                    errors.append(f"Baris {row_num}: SOAL atau GAMBAR harus diisi (minimal salah satu)")
                    continue
                
                if kunci not in ["A", "B", "C", "D", "E"]:
                    errors.append(f"Baris {row_num}: KUNCI harus A/B/C/D/E, ditemukan: {kunci}")
                    continue
                
                # cari subtest
                try:
                    subtest = Subtest.objects.get(code=kode_subtes)
                except Subtest.DoesNotExist:
                    errors.append(f"Baris {row_num}: Kode Subtes '{kode_subtes}' tidak ditemukan di database")
                    continue
                
                # Handle gambar
                soal_image = None
                if gambar_path:
                    from django.core.files import File
                    from urllib.parse import urlparse
                    import requests
                    import tempfile
                    import os
                    
                    # Cek apakah path adalah URL
                    parsed = urlparse(gambar_path)
                    if parsed.scheme in ['http', 'https']:
                        # Download dari URL
                        try:
                            response = requests.get(gambar_path, timeout=10)
                            response.raise_for_status()
                            # Simpan ke temporary file
                            img_ext = os.path.splitext(parsed.path)[1] or '.jpg'
                            with tempfile.NamedTemporaryFile(delete=False, suffix=img_ext) as img_tmp:
                                img_tmp.write(response.content)
                                img_tmp_path = img_tmp.name
                            
                            # Buka file dan simpan ke model
                            with open(img_tmp_path, 'rb') as img_file:
                                soal_image = File(img_file, name=os.path.basename(parsed.path) or f'image_{row_num}{img_ext}')
                            os.unlink(img_tmp_path)
                        except Exception as e:
                            errors.append(f"Baris {row_num}: Gagal download gambar dari URL: {str(e)}")
                    else:
                        # Path file lokal (relatif ke media/soal_images/)
                        from django.conf import settings
                        full_path = os.path.join(settings.MEDIA_ROOT, 'soal_images', gambar_path)
                        if os.path.exists(full_path):
                            with open(full_path, 'rb') as img_file:
                                soal_image = File(img_file, name=os.path.basename(gambar_path))
                        else:
                            errors.append(f"Baris {row_num}: File gambar tidak ditemukan: {gambar_path}")
                
                # buat soal baru
                soal = Soal.objects.create(
                    subtest=subtest,
                    soal_text=soal_text or "",  # Bisa kosong jika hanya gambar
                    option_a=option_a,
                    option_b=option_b,
                    option_c=option_c,
                    option_d=option_d,
                    option_e=option_e,
                    correct_answer=kunci,
                )
                
                # Set gambar jika ada
                if soal_image:
                    soal.soal_image = soal_image
                    soal.save()
                
                created += 1
                
            except Exception as e:
                errors.append(f"Baris {row_num}: {str(e)}")
        
        wb.close()
        
        return Response(
            {
                'success': True,
                'created': created,
                'errors': errors[:10],  # batasi error yang ditampilkan
                'total_errors': len(errors),
            },
            status=status.HTTP_200_OK
        )
        
    except Exception as e:
        return Response(
            {'error': f'Gagal membaca file Excel: {str(e)}'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['POST'])
def submit_jawaban(request):
    """
    POST /api/submit-jawaban/
    
    Submit jawaban user untuk suatu subtest dan hitung skor.
    
    Body: {
        "username": "user123",
        "subtest_code": "LBI",
        "batch_id": "batch-1",
        "jawaban": {"1": "A", "2": "B", "3": "C", ...},  # {soal_id: jawaban}
        "durasi_detik": 1800  # optional
    }
    """
    username = request.data.get('username')
    subtest_code = request.data.get('subtest_code', '').upper()
    batch_id = request.data.get('batch_id', '')
    jawaban = request.data.get('jawaban', {})
    durasi_detik = request.data.get('durasi_detik')
    
    import logging
    logger = logging.getLogger(__name__)
    logger.info("=" * 50)
    logger.info(f"SUBMIT JAWABAN - Username: {username}, Subtest: {subtest_code}, Batch: {batch_id}")
    logger.info(f"Received jawaban type: {type(jawaban)}, count: {len(jawaban) if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban keys (first 10): {list(jawaban.keys())[:10] if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban sample: {dict(list(jawaban.items())[:5]) if isinstance(jawaban, dict) and jawaban else 'EMPTY'}")
    
    if not username or not subtest_code or not batch_id:
        return Response(
            {'error': 'username, subtest_code, dan batch_id wajib'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    try:
        subtest = Subtest.objects.get(code=subtest_code)
    except Subtest.DoesNotExist:
        return Response(
            {'error': f'Subtest dengan code {subtest_code} tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # ambil atau buat HasilTryout
    hasil, created = HasilTryout.objects.get_or_create(
        user=user,
        subtest=subtest,
        batch_id=batch_id,
        defaults={
            'jawaban': {},
            'waktu_mulai': timezone.now(),
        }
    )
    
    # update jawaban
    # jawaban dari frontend bisa dalam format {index: "A"} atau {soal_id: "A"}
    # kita perlu convert ke format {soal_id: "A"}
    jawaban_final = {}
    
    # ambil semua soal untuk subtest ini, urutkan sesuai order di API
    soal_list = list(Soal.objects.filter(subtest=subtest).order_by('id'))
    
    # Buat mapping soal_id untuk validasi
    soal_id_set = {str(soal.id) for soal in soal_list}
    soal_id_list = [str(soal.id) for soal in soal_list]  # Untuk debugging
    
    logger.info(f"Processing jawaban for subtest {subtest_code}, batch {batch_id}")
    logger.info(f"Total soal: {len(soal_list)}")
    logger.info(f"Soal IDs in database (first 10): {soal_id_list[:10]}")
    logger.info(f"Received jawaban type: {type(jawaban)}, count: {len(jawaban) if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban keys (first 10): {list(jawaban.keys())[:10] if isinstance(jawaban, dict) else 'N/A'}")
    logger.info(f"Received jawaban sample (first 5): {dict(list(jawaban.items())[:5]) if isinstance(jawaban, dict) else 'N/A'}")
    
    if not jawaban or (isinstance(jawaban, dict) and len(jawaban) == 0):
        logger.error("WARNING: Jawaban kosong atau tidak ada!")
        return Response(
            {'error': 'Jawaban kosong. Pastikan Anda sudah mengisi jawaban sebelum submit.'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    matched_count = 0
    converted_count = 0
    skipped_count = 0
    
    # Debug: print semua soal_id yang ada di database
    logger.info(f"🔍 All soal_ids in database for {subtest_code}: {sorted([int(sid) for sid in soal_id_list])[:20]}")
    
    for key, value in jawaban.items():
        key_str = str(key).strip()
        value_str = str(value).strip().upper() if value else ""
        
        if not value_str or value_str not in ['A', 'B', 'C', 'D', 'E']:
            logger.warning(f"⚠️ Skipping invalid value for key {key_str}: '{value_str}' (must be A/B/C/D/E)")
            skipped_count += 1
            continue
        
        # Prioritas 1: Jika key_str ada di soal_id_set, langsung pakai sebagai soal_id
        if key_str in soal_id_set:
            jawaban_final[key_str] = value_str
            matched_count += 1
            logger.info(f"✅ Matched soal_id directly: {key_str} -> {value_str}")
            continue
        
        # Prioritas 2: Coba convert key sebagai integer (bisa jadi soal_id atau index)
        try:
            key_int = int(key_str)
            
            # Cek apakah key_int adalah soal_id yang valid
            if key_int in [int(sid) for sid in soal_id_list]:
                # Ini adalah soal_id yang valid, tapi dalam format integer
                jawaban_final[str(key_int)] = value_str
                matched_count += 1
                logger.info(f"✅ Matched soal_id (as int): {key_int} -> {value_str}")
                continue
            
            # Jika bukan soal_id, coba sebagai index (0-based)
            if 0 <= key_int < len(soal_list):
                # Ini adalah index, convert ke soal_id
                soal_id = str(soal_list[key_int].id)
                jawaban_final[soal_id] = value_str
                converted_count += 1
                logger.info(f"✅ Converted index {key_int} to soal_id {soal_id} -> {value_str}")
            else:
                # Key adalah angka tapi di luar range index dan bukan soal_id
                logger.warning(f"❌ Skipping key out of range: {key_str} (not a valid soal_id, max index: {len(soal_list)-1})")
                skipped_count += 1
        except (ValueError, TypeError):
            # Key bukan angka dan tidak ada di soal_id_set, skip
            logger.warning(f"❌ Skipping invalid key (not number, not in soal_id_set): {key_str}")
            skipped_count += 1
    
    logger.info(f"📊 Conversion summary: matched={matched_count}, converted={converted_count}, skipped={skipped_count}, final_count={len(jawaban_final)}")
    
    logger.info(f"Final jawaban_final: {len(jawaban_final)} entries")
    logger.info(f"Sample jawaban_final keys: {list(jawaban_final.keys())[:10]}")
    logger.info(f"Sample jawaban_final values: {list(jawaban_final.values())[:10]}")
    
    hasil.jawaban = jawaban_final
    logger.info(f"Jawaban assigned to hasil.jawaban, total: {len(hasil.jawaban)}")
    hasil.waktu_selesai = timezone.now()
    if durasi_detik:
        hasil.durasi_detik = durasi_detik
    
    # Simpan dulu sebelum hitung skor
    hasil.save(update_fields=['jawaban', 'waktu_selesai', 'durasi_detik'])
    logger.info(f"✅ HasilTryout saved with jawaban count: {len(hasil.jawaban)}")
    
    # hitung skor
    hasil.hitung_skor()
    logger.info(f"✅ Skor calculated: benar={hasil.jumlah_benar}, salah={hasil.jumlah_salah}, kosong={hasil.jumlah_kosong}, skor={hasil.skor}%")
    
    # Pastikan skor dalam range 0-100
    skor_value = float(hasil.skor) if hasil.skor is not None else 0.0
    if skor_value < 0:
        skor_value = 0.0
    elif skor_value > 100:
        skor_value = 100.0
    
    return Response(
        {
            'success': True,
            'hasil': {
                'id': hasil.id,
                'subtest_code': subtest.code,
                'subtest_nama': subtest.nama_subtest,
                'batch_id': hasil.batch_id,
                'jumlah_benar': hasil.jumlah_benar,
                'jumlah_salah': hasil.jumlah_salah,
                'jumlah_kosong': hasil.jumlah_kosong,
                'skor': round(skor_value, 2),  # Bulatkan ke 2 desimal
                'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            }
        },
        status=status.HTTP_200_OK
    )


@api_view(['GET'])
def admin_dashboard(request):
    """
    GET /api/admin/dashboard/
    
    Mengembalikan statistik dan data untuk dashboard admin.
    Hanya bisa diakses oleh admin (is_staff=True).
    
    Query param: ?username=admin (wajib)
    """
    # Validasi admin dari username
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Statistik umum
    total_users = User.objects.count()
    total_students = User.objects.filter(is_staff=False).count()
    total_admins = User.objects.filter(is_staff=True).count()
    
    # Statistik subtest
    total_subtests = Subtest.objects.count()
    total_soal = Soal.objects.count()
    
    # Statistik hasil tryout
    total_hasil = HasilTryout.objects.count()
    hasil_dengan_skor = HasilTryout.objects.exclude(skor=0).count()
    avg_skor = HasilTryout.objects.aggregate(avg=Avg('skor'))['avg'] or 0
    
    # Statistik per subtest
    subtest_stats = []
    for subtest in Subtest.objects.all():
        hasil_subtest = HasilTryout.objects.filter(subtest=subtest)
        subtest_stats.append({
            'code': subtest.code,
            'nama': subtest.nama_subtest,
            'jumlah_soal': subtest.jumlah_soal,
            'total_pengerjaan': hasil_subtest.count(),
            'avg_skor': float(hasil_subtest.aggregate(avg=Avg('skor'))['avg'] or 0),
        })
    
    # Top 10 users dengan skor tertinggi
    top_users = HasilTryout.objects.values(
        'user__username', 'user__first_name', 'user__last_name'
    ).annotate(
        total_skor=Avg('skor'),
        total_pengerjaan=Count('id')
    ).order_by('-total_skor')[:10]
    
    return Response({
        'stats': {
            'users': {
                'total': total_users,
                'students': total_students,
                'admins': total_admins,
            },
            'subtests': {
                'total': total_subtests,
                'total_soal': total_soal,
            },
            'hasil_tryout': {
                'total': total_hasil,
                'dengan_skor': hasil_dengan_skor,
                'avg_skor': round(float(avg_skor), 2),
            },
        },
        'subtest_stats': subtest_stats,
        'top_users': [
            {
                'username': u['user__username'],
                'name': f"{u['user__first_name']} {u['user__last_name']}".strip() or u['user__username'],
                'avg_skor': round(float(u['total_skor']), 2),
                'total_pengerjaan': u['total_pengerjaan'],
            }
            for u in top_users
        ],
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def admin_list_soal(request):
    """
    GET /api/admin/soal/
    
    List semua soal dengan pagination.
    Query params: ?username=admin&page=1&limit=50&subtest_code=PPU&search=...
    """
    # Validasi admin
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from django.core.paginator import Paginator
    
    subtest_code = request.GET.get('subtest_code', '').upper()
    search = request.GET.get('search', '').strip()
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    queryset = Soal.objects.select_related('subtest').all()
    
    if subtest_code:
        queryset = queryset.filter(subtest__code=subtest_code)
    
    if search:
        queryset = queryset.filter(
            Q(soal_text__icontains=search) |
            Q(option_a__icontains=search) |
            Q(option_b__icontains=search) |
            Q(option_c__icontains=search) |
            Q(option_d__icontains=search) |
            Q(option_e__icontains=search)
        )
    
    paginator = Paginator(queryset, limit)
    page_obj = paginator.get_page(page)
    
    data = []
    for soal in page_obj:
        image_url = None
        if soal.soal_image:
            request_scheme = request.scheme if hasattr(request, 'scheme') else 'http'
            request_host = request.get_host() if hasattr(request, 'get_host') else 'localhost:8000'
            image_url = f"{request_scheme}://{request_host}{soal.soal_image.url}"
        
        data.append({
            'id': soal.id,
            'subtest_code': soal.subtest.code,
            'subtest_nama': soal.subtest.nama_subtest,
            'soal_text': soal.soal_text[:200] + '...' if len(soal.soal_text) > 200 else soal.soal_text,
            'soal_image': image_url,
            'has_image': bool(soal.soal_image),
            'option_a': soal.option_a,
            'option_b': soal.option_b,
            'option_c': soal.option_c,
            'option_d': soal.option_d,
            'option_e': soal.option_e,
            'correct_answer': soal.correct_answer,
            'created_at': soal.created_at.isoformat(),
        })
    
    return Response({
        'results': data,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_prev': page_obj.has_previous(),
        },
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def admin_list_users(request):
    """
    GET /api/admin/users/
    
    List semua users dengan statistik.
    Query param: ?username=admin (wajib)
    """
    # Validasi admin
    username = request.GET.get('username', '').strip()
    if not username:
        return Response(
            {'error': 'Username wajib. Kirim sebagai query parameter: ?username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        user = User.objects.get(username=username)
        if not user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    users = User.objects.annotate(
        total_hasil=Count('hasil_tryout'),
        avg_skor=Avg('hasil_tryout__skor')
    ).order_by('-date_joined')
    
    data = []
    for user in users:
        data.append({
            'id': user.id,
            'username': user.username,
            'name': user.get_full_name() or user.username,
            'email': user.email or '',
            'is_staff': user.is_staff,
            'is_active': user.is_active,
            'date_joined': user.date_joined.isoformat(),
            'total_hasil': user.total_hasil,
            'avg_skor': round(float(user.avg_skor or 0), 2),
        })
    
    return Response(data, status=status.HTTP_200_OK)


@api_view(['GET'])
def admin_list_hasil(request):
    """
    GET /api/admin/hasil/
    
    List semua hasil tryout dengan filter.
    Query params: ?admin_username=admin&username=...&subtest_code=...&page=1&limit=50
    """
    # Validasi admin
    admin_username = request.GET.get('admin_username', '').strip()
    if not admin_username:
        return Response(
            {'error': 'admin_username wajib. Kirim sebagai query parameter: ?admin_username=admin'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        admin_user = User.objects.get(username=admin_username)
        if not admin_user.is_staff:
            return Response(
                {'error': 'Akses ditolak. Hanya admin yang bisa mengakses halaman ini.'},
                status=status.HTTP_403_FORBIDDEN
            )
    except User.DoesNotExist:
        return Response(
            {'error': 'Admin user tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    from django.core.paginator import Paginator
    
    username = request.GET.get('username', '').strip()
    subtest_code = request.GET.get('subtest_code', '').upper()
    page = int(request.GET.get('page', 1))
    limit = int(request.GET.get('limit', 50))
    
    queryset = HasilTryout.objects.select_related('user', 'subtest').all()
    
    if username:
        queryset = queryset.filter(user__username__icontains=username)
    
    if subtest_code:
        queryset = queryset.filter(subtest__code=subtest_code)
    
    queryset = queryset.order_by('-waktu_selesai', '-created_at')
    
    paginator = Paginator(queryset, limit)
    page_obj = paginator.get_page(page)
    
    data = []
    for hasil in page_obj:
        data.append({
            'id': hasil.id,
            'username': hasil.user.username,
            'user_name': hasil.user.get_full_name() or hasil.user.username,
            'subtest_code': hasil.subtest.code,
            'subtest_nama': hasil.subtest.nama_subtest,
            'batch_id': hasil.batch_id,
            'jumlah_benar': hasil.jumlah_benar,
            'jumlah_salah': hasil.jumlah_salah,
            'jumlah_kosong': hasil.jumlah_kosong,
            'skor': round(float(hasil.skor), 2),
            'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            'durasi_detik': hasil.durasi_detik,
        })
    
    return Response({
        'results': data,
        'pagination': {
            'page': page,
            'limit': limit,
            'total': paginator.count,
            'total_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
            'has_prev': page_obj.has_previous(),
        },
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
def riwayat_nilai(request, username):
    """
    GET /api/riwayat-nilai/{username}/
    
    Ambil riwayat nilai tryout per subtest untuk user tertentu.
    """
    try:
        user = User.objects.get(username=username)
    except User.DoesNotExist:
        return Response(
            {'error': 'User tidak ditemukan'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    hasil_list = HasilTryout.objects.filter(user=user).select_related('subtest').order_by('-waktu_selesai', '-created_at')
    
    data = []
    for hasil in hasil_list:
        # Pastikan skor dalam range 0-100
        skor_value = float(hasil.skor) if hasil.skor is not None else 0.0
        if skor_value < 0:
            skor_value = 0.0
        elif skor_value > 100:
            skor_value = 100.0
        
        data.append({
            'id': hasil.id,
            'batch_id': hasil.batch_id,
            'subtest_code': hasil.subtest.code,
            'subtest_nama': hasil.subtest.nama_subtest,
            'jumlah_benar': hasil.jumlah_benar,
            'jumlah_salah': hasil.jumlah_salah,
            'jumlah_kosong': hasil.jumlah_kosong,
            'skor': round(skor_value, 2),  # Bulatkan ke 2 desimal
            'waktu_selesai': hasil.waktu_selesai.isoformat() if hasil.waktu_selesai else None,
            'tanggal': hasil.waktu_selesai.strftime('%Y-%m-%d') if hasil.waktu_selesai else hasil.created_at.strftime('%Y-%m-%d'),
        })
    
    return Response(data, status=status.HTTP_200_OK)
