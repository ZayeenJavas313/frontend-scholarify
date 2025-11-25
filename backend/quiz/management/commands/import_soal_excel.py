from django.core.management.base import BaseCommand, CommandError
from quiz.models import Subtest, Soal

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled at runtime
    load_workbook = None


class Command(BaseCommand):
    """
    Import bank soal dari file Excel.

    Format kolom yang diharapkan (baris header ke-1):
      A: Kode Subtes   (mis. LBI, PU, PK, dst. -> akan dicocokkan ke Subtest.code)
      B: SOAL          (teks soal / bacaan)
      C: A             (opsi jawaban A)
      D: B
      E: C
      F: D
      G: E
      H: KUNCI         (huruf A/B/C/D/E)

    Setiap baris sesudah header akan dibuatkan 1 instance `Soal`.
    """

    help = "Import bank soal dari file Excel dengan format kolom: Kode Subtes, SOAL, A, B, C, D, E, KUNCI"

    def add_arguments(self, parser):
        parser.add_argument(
            "path",
            type=str,
            help="Path ke file Excel (.xlsx) yang akan di-import",
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError(
                "openpyxl belum ter‑install. Tambahkan 'openpyxl' ke requirements dan install dulu."
            )

        path = options["path"]
        try:
            wb = load_workbook(filename=path, read_only=True, data_only=True)
        except Exception as exc:
            raise CommandError(f"Gagal membuka file Excel: {exc}")

        ws = wb.active

        # baca header untuk memastikan format benar
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        expected = ["Kode Subtes", "SOAL", "A", "B", "C", "D", "E", "KUNCI"]
        if any(h is None for h in header) or len(header) < len(expected):
            self.stdout.write(
                self.style.WARNING(
                    "Header tidak lengkap, tetap akan diproses tapi pastikan urutan kolom sesuai: "
                    + ", ".join(expected)
                )
            )

        created = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2):  # mulai dari baris kedua
            kode = (row[0].value or "").strip() if row[0].value is not None else ""
            if not kode:
                continue  # baris kosong, lewati

            try:
                subtest = Subtest.objects.get(code=kode)
            except Subtest.DoesNotExist:
                skipped += 1
                self.stdout.write(
                    self.style.WARNING(f"Skip baris {row[0].row}: Subtest code '{kode}' tidak ditemukan.")
                )
                continue

            soal_text = (row[1].value or "").strip() if row[1].value is not None else ""
            option_a = (row[2].value or "").strip() if row[2].value is not None else ""
            option_b = (row[3].value or "").strip() if row[3].value is not None else ""
            option_c = (row[4].value or "").strip() if row[4].value is not None else ""
            option_d = (row[5].value or "").strip() if row[5].value is not None else ""
            option_e = (row[6].value or "").strip() if row[6].value is not None else ""
            kunci = (row[7].value or "").strip().upper() if len(row) > 7 and row[7].value is not None else ""

            if kunci not in dict(Soal.JAWAB_CHOICES):
                skipped += 1
                self.stdout.write(
                    self.style.WARNING(
                        f"Skip baris {row[0].row}: kunci jawaban '{kunci}' tidak valid (harus A/B/C/D/E)."
                    )
                )
                continue

            Soal.objects.create(
                subtest=subtest,
                soal_text=soal_text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                option_e=option_e,
                correct_answer=kunci,
            )
            created += 1

        self.stdout.write(
            self.style.SUCCESS(f"Import selesai. Soal dibuat: {created}, baris di-skip: {skipped}")
        )


